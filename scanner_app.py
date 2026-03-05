"""
Tarama & Barkod Okuma Uygulaması
- TWAIN tarayıcıdan otomatik tarama (arkalı önlü dahil)
- Ön + arka yüz aynı anda önizleme
- Serbest kırpma
- Barkod okuma (pyzbar) - çapraz barkod desteği, yukarıdan aşağıya sıralama
- Barkod yakınındaki metin OCR ile "Konum" kolonuna yazılır (pytesseract)
- Kolon yönetimi ve Excel aktarma
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog
from PIL import Image, ImageTk, ImageEnhance
import os

# Barkod okuma
try:
    from pyzbar.pyzbar import decode as decode_barcodes
    PYZBAR_AVAILABLE = True
except ImportError:
    PYZBAR_AVAILABLE = False

# Excel
try:
    from openpyxl import Workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# TWAIN tarayıcı
try:
    import twain
    TWAIN_AVAILABLE = True
except ImportError:
    TWAIN_AVAILABLE = False

# OCR - barkod yakınındaki metin için
try:
    import pytesseract
    TESSERACT_AVAILABLE = True
except ImportError:
    TESSERACT_AVAILABLE = False


# ─────────────────────────────────────────────
# Renk ve Stil Sabitleri
# ─────────────────────────────────────────────
BG_DARK = "#1e1e2e"
BG_PANEL = "#2a2a3e"
BG_CARD = "#33334d"
FG_TEXT = "#e0e0f0"
FG_DIM = "#8888aa"
ACCENT = "#7c5cfc"
ACCENT_HOVER = "#9b7fff"
ACCENT_RED = "#ff5572"
ACCENT_GREEN = "#50fa7b"
ACCENT_ORANGE = "#ffb86c"
BORDER_COLOR = "#444466"
CROP_COLOR = "#00ff88"
FONT_FAMILY = "Segoe UI"


class ScannerManager:
    """TWAIN tarayıcı yönetimi."""

    def __init__(self):
        self.source_manager = None
        self.source = None

    def scan(self):
        """Tarayıcıdan görüntü al. Başarısız olursa None döner."""
        if not TWAIN_AVAILABLE:
            return None
        try:
            self.source_manager = twain.SourceManager(0)
            self.source = self.source_manager.OpenSource()
            if self.source is None:
                self.close()
                return None

            self.source.SetCapability(twain.ICAP_PIXELTYPE, twain.TWTY_UINT16, twain.TWPT_RGB)
            self.source.SetCapability(twain.ICAP_XRESOLUTION, twain.TWTY_FIX32, 300.0)
            self.source.SetCapability(twain.ICAP_YRESOLUTION, twain.TWTY_FIX32, 300.0)

            self.source.RequestAcquire(0, 0)
            info = self.source.GetImageInfo()
            if info:
                handle = self.source.XferImageNatively()[0]
                image = twain.DIBToBMFile(handle)
                twain.GlobalHandleFree(handle)
                import io
                pil_image = Image.open(io.BytesIO(image))
                self.close()
                return pil_image

            self.close()
            return None
        except Exception as e:
            print(f"Tarama hatası: {e}")
            self.close()
            return None

    def close(self):
        try:
            if self.source:
                self.source.destroy()
                self.source = None
            if self.source_manager:
                self.source_manager.destroy()
                self.source_manager = None
        except:
            pass


# ─────────────────────────────────────────────
# ImagePanel - tek görüntü paneli (ön veya arka)
# ─────────────────────────────────────────────
class ImagePanel:
    """Tek görüntü paneli: canvas, zoom, kırpma, geçmiş."""

    def __init__(self, parent, label):
        self.frame = tk.Frame(parent, bg=BG_PANEL)
        self.original_image = None
        self.photo_image = None
        self.zoom_level = 1.0
        self.image_history = []
        self.crop_start = None
        self.crop_rect_id = None
        self.crop_region = None
        self._build_ui(label)

    def _build_ui(self, label):
        header = tk.Frame(self.frame, bg=BG_PANEL)
        header.pack(fill=tk.X, padx=5, pady=(6, 2))

        tk.Label(
            header, text=label,
            font=(FONT_FAMILY, 10, "bold"), bg=BG_PANEL, fg=FG_TEXT
        ).pack(side=tk.LEFT)

        self.zoom_label = tk.Label(
            header, text="—",
            font=(FONT_FAMILY, 8), bg=BG_PANEL, fg=FG_DIM
        )
        self.zoom_label.pack(side=tk.RIGHT)

        canvas_frame = tk.Frame(self.frame, bg=BORDER_COLOR, padx=1, pady=1)
        canvas_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=(0, 6))

        self.canvas = tk.Canvas(
            canvas_frame, bg="#1a1a2e", highlightthickness=0, cursor="crosshair"
        )

        h_scroll = ttk.Scrollbar(canvas_frame, orient=tk.HORIZONTAL, command=self.canvas.xview)
        v_scroll = ttk.Scrollbar(canvas_frame, orient=tk.VERTICAL, command=self.canvas.yview)
        self.canvas.configure(xscrollcommand=h_scroll.set, yscrollcommand=v_scroll.set)

        h_scroll.pack(side=tk.BOTTOM, fill=tk.X)
        v_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.canvas.pack(fill=tk.BOTH, expand=True)

        self.canvas.bind("<ButtonPress-1>", self._on_mouse_down)
        self.canvas.bind("<B1-Motion>", self._on_mouse_drag)
        self.canvas.bind("<ButtonRelease-1>", self._on_mouse_up)
        self.canvas.bind("<MouseWheel>", self._on_mouse_wheel)

    def set_image(self, pil_image, save_history=True):
        if save_history and self.original_image is not None:
            self.image_history.append(self.original_image.copy())
        self.original_image = pil_image.copy()
        self.zoom_level = 1.0
        self.crop_region = None
        self._clear_crop_rect()
        self._update_display()

    def undo(self):
        if self.image_history:
            self.original_image = self.image_history.pop()
            self.zoom_level = 1.0
            self.crop_region = None
            self._clear_crop_rect()
            self._update_display()
            return True
        return False

    def _update_display(self):
        if self.original_image is None:
            return
        w, h = self.original_image.size
        nw, nh = int(w * self.zoom_level), int(h * self.zoom_level)
        display = self.original_image.resize((nw, nh), Image.Resampling.LANCZOS)
        self.photo_image = ImageTk.PhotoImage(display)
        self.canvas.delete("all")
        self.canvas.create_image(0, 0, anchor=tk.NW, image=self.photo_image)
        self.canvas.config(scrollregion=(0, 0, nw, nh))
        self.zoom_label.config(text=f"{int(self.zoom_level * 100)}%")

    def _on_mouse_wheel(self, event):
        if self.original_image is None:
            return
        if event.delta > 0:
            self.zoom_level = min(5.0, self.zoom_level * 1.1)
        else:
            self.zoom_level = max(0.1, self.zoom_level / 1.1)
        self._update_display()

    def _canvas_coords(self, event):
        return self.canvas.canvasx(event.x), self.canvas.canvasy(event.y)

    def _on_mouse_down(self, event):
        if self.original_image is None:
            return
        self._clear_crop_rect()
        self.crop_start = self._canvas_coords(event)

    def _on_mouse_drag(self, event):
        if self.crop_start is None:
            return
        x, y = self._canvas_coords(event)
        self._clear_crop_rect()
        x0, y0 = self.crop_start
        self.crop_rect_id = self.canvas.create_rectangle(
            x0, y0, x, y, outline=CROP_COLOR, width=2, dash=(5, 3)
        )

    def _on_mouse_up(self, event):
        if self.crop_start is None:
            return
        x, y = self._canvas_coords(event)
        x0, y0 = self.crop_start
        if abs(x - x0) < 5 or abs(y - y0) < 5:
            self._clear_crop_rect()
            self.crop_start = None
            return
        w, h = self.original_image.size
        rx0 = max(0, min(int(min(x0, x) / self.zoom_level), w))
        ry0 = max(0, min(int(min(y0, y) / self.zoom_level), h))
        rx1 = max(0, min(int(max(x0, x) / self.zoom_level), w))
        ry1 = max(0, min(int(max(y0, y) / self.zoom_level), h))
        self.crop_region = (rx0, ry0, rx1, ry1)
        self.crop_start = None

    def _clear_crop_rect(self):
        if self.crop_rect_id:
            self.canvas.delete(self.crop_rect_id)
            self.crop_rect_id = None

    def get_cropped_image(self):
        if self.original_image is None:
            return None
        return self.original_image.crop(self.crop_region) if self.crop_region else self.original_image


# ─────────────────────────────────────────────
# ImageViewer - ön + arka çift panel
# ─────────────────────────────────────────────
class ImageViewer(tk.Frame):
    """Ön ve arka yüzü aynı anda gösteren çift panelli görüntü alanı."""

    def __init__(self, parent):
        super().__init__(parent, bg=BG_PANEL, highlightthickness=0)
        self._build_ui()

    def _build_ui(self):
        header = tk.Frame(self, bg=BG_PANEL)
        header.pack(fill=tk.X, padx=10, pady=(10, 4))

        tk.Label(
            header, text="Goruntu Onizleme",
            font=(FONT_FAMILY, 13, "bold"), bg=BG_PANEL, fg=FG_TEXT
        ).pack(side=tk.LEFT)

        self.info_label = tk.Label(
            self, text="Tarayicidan tara veya dosya ac",
            font=(FONT_FAMILY, 9), bg=BG_PANEL, fg=FG_DIM
        )
        self.info_label.pack(fill=tk.X, padx=10)

        paned = tk.PanedWindow(
            self, orient=tk.HORIZONTAL,
            bg=BG_DARK, sashwidth=6, sashrelief=tk.FLAT, borderwidth=0
        )
        paned.pack(fill=tk.BOTH, expand=True, padx=8, pady=(4, 8))

        self.front_panel = ImagePanel(paned, "On Yuz")
        self.back_panel = ImagePanel(paned, "Arka Yuz")

        paned.add(self.front_panel.frame, minsize=180, stretch="always")
        paned.add(self.back_panel.frame, minsize=180, stretch="always")

    # ── Uyumluluk özellikleri ──────────────────
    @property
    def original_image(self):
        return self.front_panel.original_image

    @property
    def crop_region(self):
        return self.front_panel.crop_region

    @property
    def image_history(self):
        return self.front_panel.image_history

    def _update_info(self):
        fi = self.front_panel.original_image
        bi = self.back_panel.original_image
        ft = f"On: {fi.size[0]}x{fi.size[1]}px" if fi else "On: -"
        bt = f"Arka: {bi.size[0]}x{bi.size[1]}px" if bi else "Arka: -"
        self.info_label.config(text=f"{ft}  |  {bt}  |  Fare ile kirpma yapabilirsiniz")

    def set_image(self, pil_image, save_history=True):
        """On yuze goruntu set et (geri donuk uyumluluk)."""
        self.front_panel.set_image(pil_image, save_history)
        self._update_info()

    def set_front_image(self, pil_image, save_history=True):
        self.front_panel.set_image(pil_image, save_history)
        self._update_info()

    def set_back_image(self, pil_image, save_history=True):
        self.back_panel.set_image(pil_image, save_history)
        self._update_info()

    def undo_image(self):
        return self.front_panel.undo()

    def get_cropped_image(self):
        return self.front_panel.get_cropped_image()

    def get_back_cropped_image(self):
        return self.back_panel.get_cropped_image()

    def get_full_image(self):
        return self.front_panel.original_image


# ─────────────────────────────────────────────
# DataTable
# ─────────────────────────────────────────────
class DataTable(tk.Frame):
    """Sag panelde kolon yonetimi ve veri tablosu."""

    def __init__(self, parent):
        super().__init__(parent, bg=BG_PANEL, highlightthickness=0)
        self.columns = ["Barkod", "Konum"]
        self.data_rows = []
        self._build_ui()

    def _build_ui(self):
        # Header
        header = tk.Frame(self, bg=BG_PANEL)
        header.pack(fill=tk.X, padx=10, pady=(10, 5))

        tk.Label(
            header, text="Veri Tablosu", font=(FONT_FAMILY, 13, "bold"),
            bg=BG_PANEL, fg=FG_TEXT
        ).pack(side=tk.LEFT)

        # Kolon yonetimi
        col_frame = tk.Frame(self, bg=BG_PANEL)
        col_frame.pack(fill=tk.X, padx=10, pady=(0, 5))

        self._make_btn(col_frame, "+ Kolon Ekle", self._add_column, ACCENT).pack(side=tk.LEFT, padx=(0, 5))
        self._make_btn(col_frame, "< Sola", self._move_column_left, ACCENT_ORANGE).pack(side=tk.LEFT, padx=(0, 5))
        self._make_btn(col_frame, "> Saga", self._move_column_right, ACCENT_ORANGE).pack(side=tk.LEFT, padx=(0, 5))
        self._make_btn(col_frame, "Kolon Sil", self._delete_column, ACCENT_RED).pack(side=tk.LEFT, padx=(0, 5))

        self.selected_col_var = tk.StringVar(value="Barkod")
        tk.Label(col_frame, text="Secili:", font=(FONT_FAMILY, 9), bg=BG_PANEL, fg=FG_DIM).pack(side=tk.LEFT, padx=(15, 3))
        self.col_combo = ttk.Combobox(col_frame, textvariable=self.selected_col_var, values=self.columns, state="readonly", width=15)
        self.col_combo.pack(side=tk.LEFT)

        # Treeview
        tree_frame = tk.Frame(self, bg=BORDER_COLOR, padx=1, pady=1)
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 5))

        style = ttk.Style()
        style.theme_use("clam")
        style.configure(
            "Data.Treeview",
            background=BG_CARD, foreground=FG_TEXT, fieldbackground=BG_CARD,
            rowheight=28, font=(FONT_FAMILY, 10), borderwidth=0
        )
        style.configure(
            "Data.Treeview.Heading",
            background="#000000", foreground="white",
            font=(FONT_FAMILY, 10, "bold"), borderwidth=0, relief="flat"
        )
        style.map("Data.Treeview", background=[("selected", ACCENT)], foreground=[("selected", "white")])
        style.map("Data.Treeview.Heading", background=[("active", "#222222")])

        self.tree = ttk.Treeview(
            tree_frame, style="Data.Treeview",
            columns=self.columns, show="headings", selectmode="extended"
        )
        self.tree.pack(fill=tk.BOTH, expand=True, side=tk.LEFT)

        scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=self.tree.yview)
        scrollbar.pack(fill=tk.Y, side=tk.RIGHT)
        self.tree.configure(yscrollcommand=scrollbar.set)

        self.tree.bind("<Double-1>", self._on_double_click)
        self.tree.bind("<Delete>", lambda e: self._delete_row())

        # Alt butonlar
        btn_frame = tk.Frame(self, bg=BG_PANEL)
        btn_frame.pack(fill=tk.X, padx=10, pady=(0, 10))

        self._make_btn(btn_frame, "▲ Yukari", self._move_row_up, "#555577").pack(side=tk.LEFT, padx=(0, 5))
        self._make_btn(btn_frame, "▼ Asagi", self._move_row_down, "#555577").pack(side=tk.LEFT, padx=(0, 5))
        self._make_btn(btn_frame, "Satir Sil", self._delete_row, ACCENT_RED).pack(side=tk.LEFT, padx=(0, 5))
        self._make_btn(btn_frame, "Tumunu Temizle", self._clear_all, ACCENT_RED).pack(side=tk.LEFT, padx=(0, 5))
        self._make_btn(btn_frame, "Elle Satir Ekle", self._add_manual_row, ACCENT).pack(side=tk.LEFT, padx=(0, 5))

        self.row_count_label = tk.Label(btn_frame, text="Satir: 0", font=(FONT_FAMILY, 9), bg=BG_PANEL, fg=FG_DIM)
        self.row_count_label.pack(side=tk.RIGHT)

        self._refresh_columns()

    def _make_btn(self, parent, text, command, color):
        btn = tk.Button(
            parent, text=text, command=command,
            font=(FONT_FAMILY, 9, "bold"),
            bg=color, fg="white", activebackground=color,
            relief="flat", cursor="hand2", padx=10, pady=4, borderwidth=0
        )
        btn.bind("<Enter>", lambda e, b=btn: b.config(bg=ACCENT_HOVER))
        btn.bind("<Leave>", lambda e, b=btn, c=color: b.config(bg=c))
        return btn

    def _refresh_columns(self):
        self.tree["columns"] = self.columns
        for col in self.columns:
            self.tree.heading(col, text=col, anchor=tk.CENTER)
            self.tree.column(col, anchor=tk.CENTER, width=120, minwidth=80)
        self.col_combo["values"] = self.columns
        if self.columns and self.selected_col_var.get() not in self.columns:
            self.selected_col_var.set(self.columns[0])
        self._reload_data()

    def _reload_data(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        for row_data in self.data_rows:
            self.tree.insert("", tk.END, values=[row_data.get(c, "") for c in self.columns])
        self.row_count_label.config(text=f"Satir: {len(self.data_rows)}")

    def add_barcode(self, barcode_value, barcode_type="", location=""):
        """Yeni bir barkod satiri ekle. Ayni barkod zaten varsa eklemez."""
        existing = [r.get("Barkod", "") for r in self.data_rows]
        if barcode_value in existing:
            return False
        row_data = {"Barkod": barcode_value, "Konum": location}
        if "Tur" in self.columns:
            row_data["Tur"] = barcode_type
        self.data_rows.append(row_data)
        self._reload_data()
        return True

    def _add_column(self):
        name = simpledialog.askstring("Kolon Ekle", "Yeni kolon adi:", parent=self)
        if name and name.strip():
            name = name.strip()
            if name in self.columns:
                messagebox.showwarning("Uyari", f"'{name}' kolonu zaten mevcut!", parent=self)
                return
            self.columns.append(name)
            self._refresh_columns()

    def _delete_column(self):
        col = self.selected_col_var.get()
        if col == "Barkod":
            messagebox.showwarning("Uyari", "'Barkod' kolonu silinemez!", parent=self)
            return
        if col and col in self.columns:
            if messagebox.askyesno("Kolon Sil", f"'{col}' kolonu silinsin mi?", parent=self):
                self.columns.remove(col)
                for row in self.data_rows:
                    row.pop(col, None)
                self._refresh_columns()

    def _move_column_left(self):
        col = self.selected_col_var.get()
        if col not in self.columns:
            return
        idx = self.columns.index(col)
        if idx > 0:
            self.columns[idx], self.columns[idx - 1] = self.columns[idx - 1], self.columns[idx]
            self._refresh_columns()

    def _move_column_right(self):
        col = self.selected_col_var.get()
        if col not in self.columns:
            return
        idx = self.columns.index(col)
        if idx < len(self.columns) - 1:
            self.columns[idx], self.columns[idx + 1] = self.columns[idx + 1], self.columns[idx]
            self._refresh_columns()

    def _delete_row(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showinfo("Bilgi", "Silmek icin satir secin (Ctrl+Click ile coklu secim).", parent=self)
            return
        indices = sorted([self.tree.index(item) for item in selected], reverse=True)
        for idx in indices:
            if 0 <= idx < len(self.data_rows):
                self.data_rows.pop(idx)
        self._reload_data()

    def _move_row_up(self):
        selected = self.tree.selection()
        if not selected:
            return
        idx = self.tree.index(selected[0])
        if idx > 0:
            self.data_rows[idx], self.data_rows[idx - 1] = self.data_rows[idx - 1], self.data_rows[idx]
            self._reload_data()
            new_item = self.tree.get_children()[idx - 1]
            self.tree.selection_set(new_item)
            self.tree.see(new_item)

    def _move_row_down(self):
        selected = self.tree.selection()
        if not selected:
            return
        idx = self.tree.index(selected[0])
        if idx < len(self.data_rows) - 1:
            self.data_rows[idx], self.data_rows[idx + 1] = self.data_rows[idx + 1], self.data_rows[idx]
            self._reload_data()
            new_item = self.tree.get_children()[idx + 1]
            self.tree.selection_set(new_item)
            self.tree.see(new_item)

    def _clear_all(self):
        if self.data_rows:
            if messagebox.askyesno("Temizle", "Tum veriler silinsin mi?", parent=self):
                self.data_rows.clear()
                self._reload_data()

    def _on_double_click(self, event):
        region = self.tree.identify_region(event.x, event.y)
        if region != "cell":
            return
        item = self.tree.identify_row(event.y)
        column = self.tree.identify_column(event.x)
        if not item or not column:
            return
        col_idx = int(column.replace("#", "")) - 1
        if col_idx < 0 or col_idx >= len(self.columns):
            return
        col_name = self.columns[col_idx]
        row_idx = self.tree.index(item)
        current_val = self.data_rows[row_idx].get(col_name, "")
        bbox = self.tree.bbox(item, column)
        if not bbox:
            return
        self._open_inline_edit(row_idx, col_name, current_val, bbox)

    def _open_inline_edit(self, row_idx, col_name, current_val, bbox):
        entry = tk.Entry(
            self.tree, font=(FONT_FAMILY, 10),
            bg=BG_CARD, fg=FG_TEXT, insertbackground=FG_TEXT,
            relief="solid", borderwidth=1
        )
        entry.insert(0, current_val)
        entry.select_range(0, tk.END)
        entry.place(x=bbox[0], y=bbox[1], width=bbox[2], height=bbox[3])
        entry.focus_set()

        def save(e=None):
            self.data_rows[row_idx][col_name] = entry.get()
            entry.destroy()
            self._reload_data()

        entry.bind("<Return>", save)
        entry.bind("<Escape>", lambda e: entry.destroy())
        entry.bind("<FocusOut>", save)

    def _add_manual_row(self):
        row_data = {col: "" for col in self.columns}
        self.data_rows.append(row_data)
        self._reload_data()
        children = self.tree.get_children()
        if children:
            last_item = children[-1]
            self.tree.selection_set(last_item)
            self.tree.see(last_item)
            bbox = self.tree.bbox(last_item, "#1")
            if bbox:
                self._open_inline_edit(len(self.data_rows) - 1, self.columns[0], "", bbox)

    def get_all_data(self):
        return self.columns[:], [row.copy() for row in self.data_rows]


# ─────────────────────────────────────────────
# ScannerApp
# ─────────────────────────────────────────────
class ScannerApp:
    """Ana uygulama sinifi."""

    def __init__(self, root):
        self.root = root
        self.root.title("Tarama & Barkod Okuma")
        self.root.geometry("1280x750")
        self.root.minsize(900, 600)
        self.root.configure(bg=BG_DARK)
        self.scanner = ScannerManager()
        self._build_ui()

    def _build_ui(self):
        # Toolbar
        toolbar = tk.Frame(self.root, bg=BG_PANEL, height=56)
        toolbar.pack(fill=tk.X, side=tk.TOP)
        toolbar.pack_propagate(False)

        tk.Label(
            toolbar, text="Tarama & Barkod Okuma",
            font=(FONT_FAMILY, 15, "bold"), bg=BG_PANEL, fg=FG_TEXT
        ).pack(side=tk.LEFT, padx=15)

        btn_defs = [
            ("Tara", self._scan_image, ACCENT),
            ("Arkali Onlu", self._scan_duplex, "#5c7cfa"),
            ("Dosya Ac", self._open_file, ACCENT),
            ("Geri Al", self._undo_image, "#555577"),
            ("180 Dondur", self._rotate_image, "#9b59b6"),
            ("Kirp", self._apply_crop, ACCENT_ORANGE),
            ("Barkod Oku", self._read_barcode, ACCENT_GREEN),
            ("Hepsini Oku", self._read_all_barcodes, "#1abc9c"),
            ("Excel'e Aktar", self._export_excel, "#2196F3"),
        ]

        for text, cmd, color in reversed(btn_defs):
            btn = tk.Button(
                toolbar, text=text, command=cmd,
                font=(FONT_FAMILY, 10, "bold"),
                bg=color, fg="white", activebackground=ACCENT_HOVER,
                relief="flat", cursor="hand2", padx=12, pady=6, borderwidth=0
            )
            btn.pack(side=tk.RIGHT, padx=(0, 6), pady=10)
            btn.bind("<Enter>", lambda e, b=btn: b.config(bg=ACCENT_HOVER))
            btn.bind("<Leave>", lambda e, b=btn, c=color: b.config(bg=c))

        # Ana icerik
        content = tk.PanedWindow(
            self.root, orient=tk.HORIZONTAL,
            bg=BG_DARK, sashwidth=6, sashrelief=tk.FLAT, borderwidth=0
        )
        content.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        self.image_viewer = ImageViewer(content)
        content.add(self.image_viewer, minsize=400, stretch="always")

        self.data_table = DataTable(content)
        content.add(self.data_table, minsize=350, stretch="always")

        # Durum cubugu
        status_bar = tk.Frame(self.root, bg=BG_PANEL, height=28)
        status_bar.pack(fill=tk.X, side=tk.BOTTOM)
        status_bar.pack_propagate(False)

        self.status_label = tk.Label(
            status_bar, text="Hazir  |  Taramak icin 'Tara' butonuna tiklayin",
            font=(FONT_FAMILY, 9), bg=BG_PANEL, fg=FG_DIM, anchor=tk.W
        )
        self.status_label.pack(fill=tk.X, padx=10, pady=4)

    def _set_status(self, text):
        self.status_label.config(text=text)
        self.root.update_idletasks()

    # ── Goruntu islemleri ──────────────────────

    def _scan_image(self):
        if not TWAIN_AVAILABLE:
            messagebox.showinfo("Bilgi", "TWAIN bulunamadi. 'Dosya Ac' kullanin.\n\npip install pytwain", parent=self.root)
            return
        self._set_status("Tarayiciya baglaniliyor...")
        try:
            image = self.scanner.scan()
            if image:
                self.image_viewer.set_front_image(image)
                self._set_status(f"Tarama basarili  |  {image.size[0]}x{image.size[1]} piksel")
            else:
                self._set_status("Tarama basarisiz veya iptal edildi")
                messagebox.showwarning("Uyari", "Tarama yapilamadi.", parent=self.root)
        except Exception as e:
            self._set_status(f"Tarama hatasi: {e}")
            messagebox.showerror("Hata", str(e), parent=self.root)

    def _scan_duplex(self):
        """Arkali onlu tarama: on + arka yuz, her ikisi de ayni anda gosterilir."""
        if not TWAIN_AVAILABLE:
            messagebox.showinfo("Bilgi", "TWAIN bulunamadi. pip install pytwain", parent=self.root)
            return

        # On yuz
        self._set_status("On yuz taranıyor...")
        self.root.update_idletasks()
        front_image = self.scanner.scan()
        if not front_image:
            self._set_status("On yuz taramasi basarisiz")
            messagebox.showwarning("Uyari", "On yuz taranamadi.", parent=self.root)
            return

        self.image_viewer.set_front_image(front_image)
        self._set_status("On yuz tarandı. Barkodlar okunuyor...")
        self.root.update_idletasks()

        front_count = 0
        if PYZBAR_AVAILABLE:
            front_barcodes = self._decode_with_rotations(front_image)
            for b in front_barcodes:
                data = b.data.decode("utf-8", errors="replace")
                loc = self._get_text_near_barcode(front_image, b)
                self.data_table.add_barcode(data, b.type, loc)
                front_count += 1

        self._set_status(f"On yuz: {front_count} barkod  |  Belgeyi cevirin...")

        # Kullanicidan onay
        confirmed = self._show_flip_dialog(front_count)
        if not confirmed:
            self._set_status(f"Arkali onlu iptal  |  On yuzden {front_count} barkod eklendi")
            return

        # Arka yuz
        self._set_status("Arka yuz taranıyor...")
        self.root.update_idletasks()
        back_image = self.scanner.scan()
        if not back_image:
            self._set_status(f"Arka yuz taramasi basarisiz  |  On yuzden {front_count} barkod eklendi")
            messagebox.showwarning("Uyari", "Arka yuz taranamadi.", parent=self.root)
            return

        self.image_viewer.set_back_image(back_image)
        self._set_status("Arka yuz barkodları okunuyor...")
        self.root.update_idletasks()

        back_count = 0
        if PYZBAR_AVAILABLE:
            back_barcodes = self._decode_with_rotations(back_image)
            for b in back_barcodes:
                data = b.data.decode("utf-8", errors="replace")
                loc = self._get_text_near_barcode(back_image, b)
                self.data_table.add_barcode(data, b.type, loc)
                back_count += 1

        total = front_count + back_count
        self._set_status(f"Arkali onlu tamamlandi  |  On: {front_count}  Arka: {back_count}  Toplam: {total} barkod")

    def _show_flip_dialog(self, front_count):
        """Belgeyi cevirme diyalogu. True = devam, False = iptal."""
        dialog = tk.Toplevel(self.root)
        dialog.title("Belgeyi Cevirin")
        dialog.configure(bg=BG_PANEL)
        dialog.resizable(False, False)
        dialog.grab_set()

        dialog.update_idletasks()
        pw, ph = 400, 190
        sx = self.root.winfo_x() + (self.root.winfo_width() - pw) // 2
        sy = self.root.winfo_y() + (self.root.winfo_height() - ph) // 2
        dialog.geometry(f"{pw}x{ph}+{sx}+{sy}")

        tk.Label(dialog, text="Belgeyi arka yuzu icin cevirin",
                 font=(FONT_FAMILY, 13, "bold"), bg=BG_PANEL, fg=FG_TEXT).pack(pady=(22, 6))
        tk.Label(dialog, text=f"On yuzden {front_count} barkod okundu.\nHazir oldugunuzda OK'e basin.",
                 font=(FONT_FAMILY, 10), bg=BG_PANEL, fg=FG_DIM, justify=tk.CENTER).pack(pady=4)

        result = tk.BooleanVar(value=False)

        btn_row = tk.Frame(dialog, bg=BG_PANEL)
        btn_row.pack(pady=16)

        def ok():
            result.set(True)
            dialog.destroy()

        tk.Button(btn_row, text="Arka Yuzu Tara", command=ok,
                  font=(FONT_FAMILY, 10, "bold"), bg=ACCENT_GREEN, fg="white",
                  relief="flat", padx=16, pady=6, cursor="hand2").pack(side=tk.LEFT, padx=8)
        tk.Button(btn_row, text="Iptal", command=dialog.destroy,
                  font=(FONT_FAMILY, 10, "bold"), bg=ACCENT_RED, fg="white",
                  relief="flat", padx=16, pady=6, cursor="hand2").pack(side=tk.LEFT, padx=8)

        dialog.wait_window()
        return result.get()

    def _open_file(self):
        filetypes = [
            ("Resim Dosyalari", "*.png *.jpg *.jpeg *.bmp *.tiff *.tif *.gif *.webp"),
            ("Tum Dosyalar", "*.*")
        ]
        filepath = filedialog.askopenfilename(title="Goruntu Dosyasi Ac", filetypes=filetypes, parent=self.root)
        if filepath:
            try:
                image = Image.open(filepath)
                if image.mode == "RGBA":
                    image = image.convert("RGB")
                self.image_viewer.set_front_image(image)
                self._set_status(f"Dosya yuklendi: {os.path.basename(filepath)}  |  {image.size[0]}x{image.size[1]}px")
            except Exception as e:
                messagebox.showerror("Hata", f"Dosya acilamadi:\n{e}", parent=self.root)

    def _apply_crop(self):
        if self.image_viewer.original_image is None:
            messagebox.showinfo("Bilgi", "Once bir goruntu tarayin veya acin.", parent=self.root)
            return
        if self.image_viewer.crop_region is None:
            messagebox.showinfo("Bilgi", "Kirpmak icin goruntu uzerinde fare ile dikdortgen cizin.", parent=self.root)
            return
        cropped = self.image_viewer.get_cropped_image()
        if cropped:
            self.image_viewer.set_image(cropped)
            self._set_status(f"Kirpma uygulandi  |  {cropped.size[0]}x{cropped.size[1]}px")

    def _rotate_image(self):
        if self.image_viewer.original_image is None:
            messagebox.showinfo("Bilgi", "Once bir goruntu tarayin veya acin.", parent=self.root)
            return
        rotated = self.image_viewer.original_image.rotate(180, expand=True)
        self.image_viewer.set_image(rotated)
        self._set_status(f"Goruntu 180 donduruldu  |  {rotated.size[0]}x{rotated.size[1]}px")

    def _undo_image(self):
        if self.image_viewer.original_image is None:
            messagebox.showinfo("Bilgi", "Geri alinacak goruntu yok.", parent=self.root)
            return
        if self.image_viewer.undo_image():
            img = self.image_viewer.original_image
            self._set_status(f"Geri alindi  |  {img.size[0]}x{img.size[1]}px")
        else:
            messagebox.showinfo("Bilgi", "Daha fazla geri alinacak islem yok.", parent=self.root)

    # ── Barkod okuma ──────────────────────────

    def _check_pyzbar(self):
        if not PYZBAR_AVAILABLE:
            messagebox.showerror("Hata", "pyzbar bulunamadi!\n\npip install pyzbar", parent=self.root)
            return False
        if self.image_viewer.original_image is None:
            messagebox.showinfo("Bilgi", "Once bir goruntu tarayin veya acin.", parent=self.root)
            return False
        return True

    def _decode_with_rotations(self, image):
        """
        Capraz/egik barkodlar icin farkli acilarla okuma dener.
        Sonuclari yukaridan asagiya (rect.top) gore siralar.
        """
        seen_data = set()
        all_barcodes = []

        def try_decode(img):
            results = []
            try:
                for b in decode_barcodes(img):
                    d = b.data.decode("utf-8", errors="replace")
                    if d not in seen_data:
                        seen_data.add(d)
                        results.append(b)
            except Exception:
                pass
            return results

        def add_sorted(results):
            results.sort(key=lambda b: b.rect.top)
            all_barcodes.extend(results)

        # 1. Orijinal goruntu
        r = try_decode(image)
        if r:
            add_sorted(r)
            return all_barcodes

        # 2. Gri tonlama
        gray = image.convert("L")
        r = try_decode(gray)
        if r:
            add_sorted(r)
            return all_barcodes

        # 3. Yuksek kontrast
        r = try_decode(ImageEnhance.Contrast(gray).enhance(2.5))
        if r:
            add_sorted(r)
            return all_barcodes

        # 4. Standart dondurmeler
        for angle in [90, 270, 180]:
            fill = (255, 255, 255) if image.mode == "RGB" else 255
            rotated = image.rotate(angle, expand=True, fillcolor=fill)
            r = try_decode(rotated)
            if not r:
                r = try_decode(rotated.convert("L"))
            if r:
                add_sorted(r)
                return all_barcodes

        # 5. Kucuk acili dondurmeler (capraz barkodlar)
        for angle in [-5, 5, -10, 10, -15, 15, -20, 20, -25, 25, -30, 30]:
            fill = (255, 255, 255) if image.mode == "RGB" else 255
            rotated = image.rotate(angle, expand=True, fillcolor=fill)
            r = try_decode(rotated)
            if r:
                add_sorted(r)
                return all_barcodes

        return all_barcodes

    def _get_text_near_barcode(self, image, barcode):
        """
        Barkodun yakinindaki metni OCR ile okur (Turkce + Ingilizce).
        pytesseract yuklu degilse bos string doner.
        """
        if not TESSERACT_AVAILABLE:
            return ""
        try:
            r = barcode.rect
            # Barkodun ustu, alti ve yanlarinda metin ara
            pad_y = max(r.height * 2, 100)
            pad_x = max(r.width // 2, 60)
            x0 = max(0, r.left - pad_x)
            y0 = max(0, r.top - pad_y)
            x1 = min(image.width, r.left + r.width + pad_x)
            y1 = min(image.height, r.top + r.height + pad_y)

            region = image.crop((x0, y0, x1, y1))
            # Gri tonlama + kontrast - OCR kalitesini arttirir
            region = region.convert("L")
            region = ImageEnhance.Contrast(region).enhance(2.0)
            # OCR icin 2x buyut
            region = region.resize((region.width * 2, region.height * 2), Image.Resampling.LANCZOS)

            # Once Turkce deneyin, yoksa sadece eng
            try:
                text = pytesseract.image_to_string(region, lang="tur+eng", config="--psm 6")
            except Exception:
                text = pytesseract.image_to_string(region, lang="eng", config="--psm 6")

            # Bos satirlari kaldir, temizle
            lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
            text = " | ".join(lines)
            return text[:200]
        except Exception:
            return ""

    def _read_barcode(self):
        """En kisa barkodu oku (on yuzden)."""
        if not self._check_pyzbar():
            return
        self._set_status("Barkod aranıyor...")
        self.root.update_idletasks()
        image_to_scan = self.image_viewer.get_cropped_image()
        try:
            barcodes = self._decode_with_rotations(image_to_scan)
            if barcodes:
                shortest = min(barcodes, key=lambda b: len(b.data))
                data = shortest.data.decode("utf-8", errors="replace")
                loc = self._get_text_near_barcode(image_to_scan, shortest)
                self.data_table.add_barcode(data, shortest.type, loc)
                self._set_status(f"En kisa barkod eklendi: {data}  |  Toplam {len(barcodes)} barkod bulundu")
            else:
                self._set_status("Goruntude barkod bulunamadi")
                messagebox.showinfo("Bilgi", "Barkod bulunamadi.\n\n• Barkodun net oldugunu kontrol edin\n• Kirparak tekrar deneyin", parent=self.root)
        except Exception as e:
            self._set_status(f"Barkod okuma hatasi: {e}")
            messagebox.showerror("Hata", str(e), parent=self.root)

    def _read_all_barcodes(self):
        """
        Tum barkodlari oku - on ve arka yuzden, yukaridan asagiya sirali.
        Barkod yakinindaki metin Konum kolonuna yazilir (pytesseract gerekli).
        """
        if not self._check_pyzbar():
            return

        if not TESSERACT_AVAILABLE:
            messagebox.showwarning(
                "Konum Okuma Devre Disi",
                "pytesseract yuklu olmadigi icin barkod yakinindaki metin Konum kolonuna yazılamaz.\n\n"
                "Kurmak icin:\n"
                "1. pip install pytesseract\n"
                "2. https://github.com/UB-Mannheim/tesseract/wiki adresinden Tesseract'i indirin\n\n"
                "Barkodlar yine de okunacak, sadece Konum bos kalacak.",
                parent=self.root
            )

        self._set_status("Tum barkodlar aranıyor (on + arka yuz)...")
        self.root.update_idletasks()

        total = 0

        # On yuz
        front_img = self.image_viewer.get_cropped_image()
        if front_img:
            front_barcodes = self._decode_with_rotations(front_img)
            for b in front_barcodes:
                data = b.data.decode("utf-8", errors="replace")
                loc = self._get_text_near_barcode(front_img, b)
                self.data_table.add_barcode(data, b.type, loc)
                total += 1

        # Arka yuz (varsa)
        back_img = self.image_viewer.get_back_cropped_image()
        if back_img:
            back_barcodes = self._decode_with_rotations(back_img)
            for b in back_barcodes:
                data = b.data.decode("utf-8", errors="replace")
                loc = self._get_text_near_barcode(back_img, b)
                self.data_table.add_barcode(data, b.type, loc)
                total += 1

        if total:
            ocr_note = " (Konum icin pytesseract gerekli)" if not TESSERACT_AVAILABLE else ""
            self._set_status(f"{total} barkod bulundu ve tabloya eklendi{ocr_note}")
        else:
            self._set_status("Goruntude barkod bulunamadi")
            messagebox.showinfo(
                "Bilgi",
                "Barkod bulunamadi.\n\n• Barkodun net oldugunu kontrol edin\n• Kirparak tekrar deneyin\n• Tarama cozunurlugunu artirin",
                parent=self.root
            )

    # ── Excel aktarma ──────────────────────────

    def _export_excel(self):
        if not OPENPYXL_AVAILABLE:
            messagebox.showerror("Hata", "openpyxl bulunamadi!\npip install openpyxl", parent=self.root)
            return

        columns, data_rows = self.data_table.get_all_data()
        if not data_rows:
            messagebox.showinfo("Bilgi", "Aktarilacak veri yok.", parent=self.root)
            return

        filepath = filedialog.asksaveasfilename(
            title="Excel Dosyasi Kaydet",
            defaultextension=".xlsx",
            filetypes=[("Excel Dosyasi", "*.xlsx")],
            initialfile="barkod_verileri.xlsx",
            parent=self.root
        )
        if not filepath:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Barkod Verileri"

            from openpyxl.styles import Font, PatternFill, Alignment

            header_font = Font(name="Segoe UI", bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
            header_align = Alignment(horizontal="center", vertical="center")

            for col_idx, col_name in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align

            data_font = Font(name="Segoe UI", size=10)
            data_align = Alignment(horizontal="center", vertical="center")

            for row_idx, row_data in enumerate(data_rows, 2):
                for col_idx, col_name in enumerate(columns, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(col_name, ""))
                    cell.font = data_font
                    cell.alignment = data_align

            for col_idx, col_name in enumerate(columns, 1):
                max_len = max([len(col_name)] + [len(str(r.get(col_name, ""))) for r in data_rows])
                col_letter = chr(64 + col_idx) if col_idx <= 26 else "A"
                ws.column_dimensions[col_letter].width = max_len + 4

            wb.save(filepath)
            self._set_status(f"Excel kaydedildi: {os.path.basename(filepath)}")
            messagebox.showinfo("Basarili", f"Excel kaydedildi:\n{filepath}", parent=self.root)

        except Exception as e:
            messagebox.showerror("Hata", f"Excel aktarma hatasi:\n{e}", parent=self.root)


def main():
    root = tk.Tk()
    try:
        from ctypes import windll
        windll.shcore.SetProcessDpiAwareness(1)
    except:
        pass
    ScannerApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
